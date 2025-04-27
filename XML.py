import xml.etree.ElementTree as ET
from openpyxl import Workbook
import os
from openpyxl.styles import Alignment

# Thư mục chứa các file XML
input_folder = '/Users/admin/Library/Mobile Documents/com~apple~CloudDocs/Tải về Iphone/GTGT Lemolove'
output_file = '/Users/admin/Downloads/ket_qua.xlsx'

# Hàm xử lý từng file XML
def process_file(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()

    data_ket_qua = []
    for elem in root.iter():
        tag_name = elem.tag.split('}')[-1]  # Bỏ namespace
        gia_tri = elem.text.strip() if elem.text else ""
        if gia_tri and tag_name.startswith('ct'):
            num_part = tag_name[2:]  # Lấy phần sau 'ct'
            if num_part and (num_part[0].isdigit()):  # Kiểm tra nếu bắt đầu bằng số
                chi_so = int(''.join(filter(str.isdigit, num_part)))  # Lấy phần số
                if 1 <= chi_so <= 43:
                    # Định dạng giá trị với phân tách hàng nghìn
                    formatted_value = f"{int(gia_tri):,}" if gia_tri.isdigit() else gia_tri
                    data_ket_qua.append({'ChiTieu': tag_name, 'GiaTri': formatted_value})
    return data_ket_qua

# Tạo workbook để ghi dữ liệu
wb = Workbook()
ws = wb.active
ws.title = "Kết Quả"

# Ghi tiêu đề
ws.append(["STT", "Tên chỉ tiêu", "Giá trị", "Tên file", "Đường dẫn"])

# Duyệt qua tất cả các file XML trong thư mục
stt = 1
for file_name in os.listdir(input_folder):
    if file_name.endswith('.xml'):
        file_path = os.path.join(input_folder, file_name)
        print(f"Đang xử lý file: {file_path}")
        data_ket_qua = process_file(file_path)

        # Ghi dữ liệu vào Excel
        for row in data_ket_qua:
            ws.append([stt, row['ChiTieu'], row['GiaTri'], file_name, file_path])
            stt += 1

        # In dữ liệu ra màn hình
        for row in data_ket_qua:
            print(f"ChiTieu: {row['ChiTieu']}, GiaTri: {row['GiaTri']}")

# Định dạng các ô giá trị là dạng số
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        try:
            cell.value = float(cell.value.replace(',', ''))  # Chuyển giá trị về dạng số
            cell.number_format = '#,##0'  # Định dạng số với phân tách hàng nghìn
            cell.alignment = Alignment(horizontal='right')  # Căn phải
        except (ValueError, AttributeError):
            pass

# Lưu file Excel
wb.save(output_file)
print(f"Dữ liệu đã được lưu vào: {output_file}")

